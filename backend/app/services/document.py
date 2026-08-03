"""Document processing service: parsing, cleaning, chunking, embedding generation."""

from __future__ import annotations

import io
import csv
import pypdf
import docx
import openpyxl
from langchain.text_splitter import RecursiveCharacterTextSplitter
from openai import AsyncOpenAI

from app.core.config import get_settings
from app.core.exceptions import UnsupportedFileTypeError, DocumentProcessingError

settings = get_settings()


class DocumentProcessor:
    """Handles multi-format document parsing, text extraction, chunking, and embedding."""

    def __init__(self) -> None:
        self.openai_client = AsyncOpenAI(api_key=settings.openai.api_key)
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=settings.rag.chunk_size,
            chunk_overlap=settings.rag.chunk_overlap,
            separators=["\n\n", "\n", " ", ""],
        )

    def extract_text(self, file_bytes: bytes, file_type: str) -> str:
        """Extract clean text from PDF, DOCX, XLSX, CSV, or TXT file."""
        ft = file_type.lower()
        if ft == "txt":
            return file_bytes.decode("utf-8", errors="ignore")
        elif ft == "pdf":
            return self._extract_pdf(file_bytes)
        elif ft == "docx":
            return self._extract_docx(file_bytes)
        elif ft == "xlsx":
            return self._extract_xlsx(file_bytes)
        elif ft == "csv":
            return self._extract_csv(file_bytes)
        else:
            raise UnsupportedFileTypeError(message=f"File format '{file_type}' is not supported.")

    def _extract_pdf(self, file_bytes: bytes) -> str:
        try:
            reader = pypdf.PdfReader(io.BytesIO(file_bytes))
            pages_text = [page.extract_text() or "" for page in reader.pages]
            return "\n\n".join(pages_text)
        except Exception as exc:
            raise DocumentProcessingError(message=f"Failed to parse PDF file: {exc}") from exc

    def _extract_docx(self, file_bytes: bytes) -> str:
        try:
            doc = docx.Document(io.BytesIO(file_bytes))
            return "\n\n".join([p.text for p in doc.paragraphs if p.text.strip()])
        except Exception as exc:
            raise DocumentProcessingError(message=f"Failed to parse DOCX file: {exc}") from exc

    def _extract_xlsx(self, file_bytes: bytes) -> str:
        try:
            wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
            text_lines = []
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                text_lines.append(f"Sheet: {sheetname}")
                for row in ws.iter_rows(values_only=True):
                    row_str = ", ".join([str(val) for val in row if val is not None])
                    if row_str.strip():
                        text_lines.append(row_str)
            return "\n".join(text_lines)
        except Exception as exc:
            raise DocumentProcessingError(message=f"Failed to parse XLSX file: {exc}") from exc

    def _extract_csv(self, file_bytes: bytes) -> str:
        try:
            content = file_bytes.decode("utf-8", errors="ignore")
            reader = csv.reader(io.StringIO(content))
            lines = [", ".join(row) for row in reader if any(cell.strip() for cell in row)]
            return "\n".join(lines)
        except Exception as exc:
            raise DocumentProcessingError(message=f"Failed to parse CSV file: {exc}") from exc

    def chunk_text(self, text: str) -> list[str]:
        """Split text into semantic chunks."""
        return self.text_splitter.split_text(text)

    async def generate_embeddings(self, texts: list[str]) -> list[list[float]]:
        """Generate OpenAI embeddings for a list of text chunks."""
        if not texts:
            return []
        try:
            response = await self.openai_client.embeddings.create(
                model=settings.openai.embedding_model,
                input=texts,
            )
            return [data.embedding for data in response.data]
        except Exception as exc:
            raise DocumentProcessingError(message=f"Embedding generation failed: {exc}") from exc
